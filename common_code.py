import configparser
from openpyxl import load_workbook,Workbook
from wordcloud import WordCloud
import matplotlib.pyplot as plt


def find_dimensions(wb,sheets,header_rows = 3,header_columns = 2): # common
    i = 0
    j = 0
    dimensions = {}
    for ws in sheets :
        i = 0
        j = 0
        while wb[ws].cell(row=header_rows+i, column=header_columns-1).value != None :
            i += 1
        while wb[ws].cell(row=header_rows-1, column=j+header_columns).value != None :
            j += 1
        dimensions[ws] = [i,j]
    return dimensions



def read_questions() :
    config = configparser.ConfigParser()
    config.read('config.ini')

    header_rows = int(config['excel_sheet_params']['header_rows'])
    header_columns = int(config['excel_sheet_params']['header_columns'])
    input_file = config['excel_sheet_params']['filename']
    output_file_name = config['excel_sheet_params']['output_xlsx_file']
    questions = []
    wb = load_workbook(filename = input_file)
    sheets = wb.get_sheet_names()
    dimensions = find_dimensions(wb,sheets,header_rows,header_columns)
    max_questions = max(dimensions.keys(), key = lambda k: dimensions[k][1])
    for i in range(header_columns,dimensions[max_questions][1]+header_columns):
        questions.append([])
        for ws in sheets:
            if dimensions[ws][1] > i-header_columns :
                for j in range(header_rows,dimensions[ws][0]+header_rows):
                    questions[i-header_columns].append(wb[ws].cell(row=j, column=i).value)

    return questions,dimensions[max_questions][1],dimensions,sheets

def generate_word_cloud(sol,filename,thread_message,eligible_qts = []) :
    if eligible_qts == [] :
        eligible_qts = [i for i in range(len(sol))]
    for i in range(len(sol)):
        wordcloud = WordCloud().generate(sol[i])
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis("off")
        plt.title('Word Cloud of Question '+ str(i+1), fontsize=14, fontweight='bold')
        plt.savefig(filename + str(eligible_qts[i]+1) +".png")
        plt.gcf().clear()

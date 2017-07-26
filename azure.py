import requests
import json
import re
import sys
from openpyxl import load_workbook,Workbook
import os.path
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import asyncio
import aiohttp
import time
from nltk.corpus import wordnet
import configparser
import nltk
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer






def output_exel(info,max_questions,dimensions,sheets,_type,thread_message):
    config = configparser.ConfigParser()
    config.read('config.ini')
    input_file = config['excel_sheet_params']['filename']
    output_file_name = config['excel_sheet_params']['output_xlsx_file']

    if os.path.isfile(output_file_name) is True :
        wb = load_workbook(filename = output_file_name)
        sheets = wb.get_sheet_names()
    else :
        wb1 = load_workbook(filename = input_file)
        wb = Workbook()
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        for sheet in sheets:
            wb.create_sheet(sheet)
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            ws1 = wb1.get_sheet_by_name(sheet.title)
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column = dimensions[sheet.title][1]+1)
            ws1 = wb1.get_sheet_by_name(sheet.title)

            for i in range(2,dimensions[sheet.title][1]+2) :
                ws.cell(row = 2,column = i,value = str(ws1.cell(row = 2, column = i).value))
            for i in range(1,dimensions[sheet.title][0]+3) :
                ws.cell(row = i,column = 1,value = str(ws1.cell(row = i, column = 1).value))

    if _type is 'sentiment' :
        key_value = 'score'
        tag_string = "sentiment analysis score : "
    if _type is 'keyword' :
        key_value = 'keyPhrases'
        tag_string = 'key words : '
    for j in range(max_questions):
        ans_no = 0
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            for i in range(dimensions[sheet.title][0]) :
                if dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value != None:
                    ws.cell(row = i+3, column = j+2, value = str(ws.cell(row = i+3, column = j+2).value) + str("\n") + tag_string + str(info[j][ans_no][key_value]))
                    ans_no += 1
                elif dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value == None:
                    ws.cell(row = i+3, column = j+2, value = str(tag_string + str(info[j][ans_no][key_value])))
                    ans_no += 1

    wb.save(output_file_name)

def output_exel_with_question(info,max_questions,dimensions,sheets,_type,thread_message):
    config = configparser.ConfigParser()
    config.read('config.ini')
    input_file = config['excel_sheet_params']['filename']
    output_file_name = config['excel_sheet_params']['output_xlsx_file']
    header_rows = int(config['excel_sheet_params']['header_rows'])
    header_columns = int(config['excel_sheet_params']['header_columns'])

    if os.path.isfile(output_file_name) is True :
        wb = load_workbook(filename = output_file_name)
        sheets = wb.get_sheet_names()
    else :
        wb = load_workbook(filename = input_file)
        wb.save(output_file_name)
        wb = load_workbook(filename = output_file_name)

    if _type is 'sentiment' :
        key_value = 'score'
        tag_string = "sentiment analysis score : "
    if _type is 'keyword' :
        key_value = 'keyPhrases'
        tag_string = 'key words : '
    for j in range(max_questions):
        ans_no = 0
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            for i in range(dimensions[sheet.title][0]) :
                if dimensions[sheet.title][1] > j and ws.cell(row = i+header_rows, column = j+header_columns).value != None:
                    ws.cell(row = i+header_rows, column = j+header_columns, value = str(ws.cell(row = i+header_rows, column = j+header_columns).value) + str("\n") + tag_string + str(info[j][ans_no][key_value]))
                    ans_no += 1
                elif dimensions[sheet.title][1] > j and ws.cell(row = i+header_rows, column = j+header_columns).value == None:
                    ws.cell(row = i+header_rows, column = j+header_columns, value = str(tag_string + str(info[j][ans_no][key_value])))
                    ans_no += 1

    wb.save(output_file_name)



def generate_piechart_for_azure_sentiment(sentiment,thread_message) :
    for i in range(len(sentiment)) :
        a = [0,0,0]
        for sent in sentiment[i] :
            if sent['score'] > .45 and sent['score'] < .55 :
                a[2] += 1
            elif sent['score'] > .5 :
                a[0] += 1
            else :
                a[1] += 1

        labels = 'positive', 'negative','neutral'
        explode = (0, 0.1, 0.2)  # only "explode" the 2nd slice (i.e. 'Hogs')
        fig1, ax1 = plt.subplots()
        ax1.pie(a, explode=explode, labels=labels, autopct='%1.1f%%',shadow=True, startangle=90)
        ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        plt.title('Sentiment Analysis for question '+str(i+1), fontsize=14, fontweight='bold')
        plt.savefig('azure_sentiment' + str(i+1) +".png")
        plt.gcf().clear()

#------------------------------file operations ends here -----------------------------------------------------------------------------------------------------------------------------------------------------------------------
@asyncio.coroutine
def call_url(url,text_data,key,i,thread_message):
    response = yield from aiohttp.ClientSession().post(url, data = text_data.encode('utf-8'), headers = {'Ocp-Apim-Subscription-Key': key, 'Content-Type':'application/json'}, params = {'numberOfLanguagesToDetect':'1'})
    response = yield from response.json()
    return [response,url,i]


def send_nlp_request_azure(url,text_data,key,thread_message):
    async_res,res_info,temp_data = [],[],[]
    sentiment,keywords = [[]] * len(text_data),[[]] * len(text_data)
    config = configparser.ConfigParser()
    config.read('config.ini')
    size_of_file = int(config['azure_params']['size_of_file'])
    try:
        for i in range(len(text_data)) :
            length = len(text_data[i]['documents'])
            temp_data = {}
            temp_data['documents'] = []

            for j in range(len(text_data[i]['documents'])):
                temp_data['documents'].append(text_data[i]['documents'][j])
                if (sys.getsizeof(temp_data['documents']) > size_of_file) or (length-1 == j) :
                    async_res.append(call_url(url[0],json.dumps(temp_data),key,i,thread_message))
                    async_res.append(call_url(url[1],json.dumps(temp_data),key,i,thread_message))
                    temp_data = {}
                    temp_data['documents'] = []



        loop = asyncio.get_event_loop()
        res_info.append(loop.run_until_complete(asyncio.wait(async_res)))
        for i in range(len(res_info[0][0])) :
            val = res_info[0][0].pop().result()

            if val[1] == url[0] :
                sentiment[val[2]] = list(sentiment[val[2]]) + val[0]['documents']
            elif val[1] == url[1] :
                keywords[val[2]] = keywords[val[2]] + val[0]['documents']

    except requests.exceptions.Timeout as e:
        print(thread_message,e)
        sys.exit(1)
    except requests.exceptions.TooManyRedirects as e:
        print(thread_message,e)
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(thread_message,e)
        sys.exit(1)

    return sentiment,keywords


#--------------------------------sentiment&keyword extraction functions end here----------------------------------------------------------------------------------------------------------------------

@asyncio.coroutine
def call_topic_url(url,header,i,j,thread_message):
    response = yield from aiohttp.ClientSession().get(url, headers = header, params = {'numberOfLanguagesToDetect':'1'})
    response = yield from response.json()
    return [response,i,j]


def send_grouping_azure_req(text_data,url,key,thread_message) :

    async_res,async_req = [],[]
    res_url,res_data = [],[]
    no_of_processes = 0
    if text_data != [] :
        for i in range(len(text_data)) :
            res_url.append([])
            res_data.append([])


        #resp_url = []
        no_of_questons = len(text_data)

        for i in range(len(text_data)) :
            length = len(text_data[i])

            temp_data = {}
            temp_data['documents'],temp_data['stopWords'],temp_data['stopPhrases'] = [],[],[]
            for j in range(len(text_data[i])):
                if text_data[i][j] is None or text_data[i][j].lower() is 'yes' :
                    text_data[i][j] = "yes, i agree with that"
                element = {}
                element['text'] = text_data[i][j]
                element['id'] = j
                temp_data['documents'].append(element)

                if (sys.getsizeof(temp_data['documents']) > 500000) or (length-1 == j) :
                    no_of_processes += 1
                    temp_data['stopWords'] += ['professional','Wipro', 'excellence']
                    temp_data['stopPhrases'] += ['professional excellence']
                    print('sent request : ', no_of_processes )
                    resp = requests.post(url, data = json.dumps(temp_data).encode('utf-8'), headers = {'Ocp-Apim-Subscription-Key': key, 'Content-Type':'application/json'}, params = {'numberOfLanguagesToDetect':'1'})
                    #print(thread_message,resp.headers)
                    print("got a response URL : ", no_of_processes)
                    if i != len(text_data)-1 :
                        time.sleep(61)
                    resp = resp.headers['operation-location']
                    res_url[i] += [resp]
                    temp_data = {}
                    temp_data['documents'],temp_data['stopWords'],temp_data['stopPhrases'] = [],[],[]




        #if 'operation-location' in res_info.headers :
            #if res_info.status_code == 202:
            #    stat_url = res_info.headers['Operation-Location']

        not_ready = True
        while not_ready:
            proc_stat = get_operation_stat(res_url,key,thread_message)
            flag = 0
            completed_processes = 0
            for i in proc_stat :
                if i[0]['status'] == 'Succeeded' :
                    res_url[i[1]][i[2]] = 'completed'
                    res_data[i[1]] += i[0]['operationProcessingResult']['topics']
                else :
                    flag += 1

            print(thread_message," : ",no_of_processes-flag ,"Requests completed out of ", no_of_processes)
            if (flag != 0):
                time.sleep(60)
            else:
                not_ready = False


    #print(thread_message," : ",res_data)
    return res_data
    #print(thread_message," : ",res_asignment)




def get_operation_stat(stat_url,api_key,thread_message):

    process_status,async_res,response = [],[],[]
    header = {'Ocp-Apim-Subscription-Key':api_key, \
            'Content-type': 'application/json',\
            'Accept': 'application/json'}
    for i in range(len(stat_url)) :
        for j in range(len(stat_url[i])) :
            if stat_url[i][j] != 'completed' :
                async_res.append(call_topic_url(stat_url[i][j],header,i,j,thread_message))

    loop = asyncio.get_event_loop()
    response.append(loop.run_until_complete(asyncio.wait(async_res)))

    for i in range(len(response[0][0])) :
        val = response[0][0].pop().result()
        process_status.append(val)

    return process_status





#-------------------------------grouping topics starts here ------------------------------------------------------------------------------------------------------------------------------------------
def azure_error_message(res_info,thread_message):
    print (thread_message," : ",'Inner Erro Message : ' + res_info['innerError']['message'])
    print (thread_message," : ",'Inner Erro Code : ' + res_info['innerError']['code'])
    print (thread_message," : ",'Over All Erro Message : ' + res_info['message'])
    print (thread_message," : ",'Over All Error Message : ' + res_info['code'])
    sys.exit(1)





#----------------------------------code for including keywords in question ----------------------------------------------------------------------------------------------------------------------------------------
def question_related_keys_extraction(keywords_info,question,gradient,treshold,thread_message) :
    question_rel,answer,q_len,question_wordcloud,a_len,extra_keys = [],[],[],[],[],[]


    for i in range(len(question)) :
        for j in range(len(question[i])) :
            question[i][j] = wordnet.synsets(question[i][j])[0]


    for i in range(len(keywords_info)) :
        extra_keys.append([])
        answer.append([])
        for word1 in keywords_info[i]:
            word1 = word1.split(" ")
            for word2 in word1 :
                count = 0
                wordFromList1 = wordnet.synsets(word2)
                if len(wordFromList1) != 0 :
                    index = -1
                    for l,val in enumerate(wordFromList1) :
                        if val.lemma_names()[0] == word2 :
                            index = l

                            break
                    if index != -1 :
                        for word3 in question[i]:
                            if wordFromList1[index] and word3:
                                s = wordFromList1[index].wup_similarity(word3)
                                if s != None and count < s :
                                    count = s
                        answer[i].append((wordFromList1[index],count))
                    else :
                        extra_keys[i].append(word2)
        q_len.append(len(question[i]))
        a_len.append(len(answer[i]))
    for i in range(len(keywords_info)) :
        if len(answer[i]) > gradient :
            while a_len[i]*(1-treshold) < len(answer[i]) :

                sorted(answer,key = lambda x: x[1])
                for j in range(gradient) :
                    question[i].append(answer[i].pop(0)[0])
                for j in range(len(answer[i])) :
                    count = answer[i][j][1]
                    for k in range(q_len[i],q_len[i]+gradient) :
                        s = answer[i][j][0].wup_similarity(question[i][k])
                        if s != None and count < s :
                            count = s
                    temp = list(answer[i][j])
                    temp[1] = count
                    answer[i][j] = tuple(temp)
                q_len[i] = len(question[i])

            string = ''
            for j in range(len(question[i])):
                question[i][j] = question[i][j].lemma_names()[0]
            for word in question[i] :
                string = string + word + " "
            question_wordcloud.append(string)

    return question_wordcloud,extra_keys

#--------------------------------------------------ectracting most relevant answers-----------------------------------------
def extract_most_relevant(data,keywords,thread_message) :
    req = []
    ps = PorterStemmer()
    for j in range(len(data)):
        lst= []
        req.append([])
        for i,x in enumerate(data[j]) :
            if data[j][i] != None :
                lst.append([nltk.pos_tag(word_tokenize(data[j][i])),i])
            else :
                lst.append([('none','NN'),i])

        for i,rev in enumerate(lst) :
            dt_tags = [t for t in rev[0] if t[1] == "VB" or t[1] == 'NN' or t[1] == 'NNP' or t[1] == 'VBG' or t[1] == 'VBP' or t[1] == 'VBZ']
            lst[i][0] = dt_tags

        for i,rev in enumerate(lst) :
            for k,val in enumerate(rev[0]) :
                temp = wordnet.synsets(lst[i][0][k][0])
                if temp != [] :
                    lst[i][0][k] = list(lst[i][0][k])
                #lst[i][0][k][0] = ps.stem(lst[i][0][k][0])

                    lst[i][0][k][0] = temp
                    lst[i][0][k] = tuple(lst[i][0][k])
                else :
                    lst[i][0].pop(j)
        #lst = [ps.stem(word[0]) for word in lst]

        stems = [ps.stem(word) for word in keywords[j]]
        synonyms = []
        for word in keywords[j] :
            for syn in wordnet.synsets(word):
                for l in syn.lemmas():
                    synonyms.append(l.name())

        #print(synonyms,'\n',stems,'\n',keywords[j],'\n')
        synonyms = [ps.stem(word) for word in synonyms]
        for i,review in enumerate(lst) :
            score = 0
            for k,words in enumerate(review[0]) :
                if words[0] in keywords[j] :
                    score += 100
                elif words[0] in stems :
                    score += 90
                elif words[0] in synonyms :
                    score += 65
            #if len(lst[i][0]) != 0 :
                #score /= len(lst[i][0])
            #else :
                #score = 0
            lst[i].append(score)
            req[j].append((score,i))

        req[j] = sorted(req[j],key = lambda x : float(x[0]),reverse = True)

    return req

def write_to_file(most_rel,data,thread_message) :

    try:
        os.remove('most_relevant_answers.csv')
    except OSError:
        pass

    f = open("most_relevant_answers.csv","a+")

    for i,question in enumerate(most_rel) :
        temp = 'question' + str(i) + ','
        f.write(temp)
        for val in question :
            if val[0] != 0 :
                f.write(data[i][val[1]]+',')
        f.write('\n')
    f.close()

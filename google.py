import requests
import json
import re
import sys
from openpyxl import load_workbook,Workbook
import os.path
from wordcloud import WordCloud
import matplotlib as mpl
mpl.use('Agg')
import argparse
from googleapiclient import discovery
from oauth2client.client import GoogleCredentials
import matplotlib.pyplot as plt
import asyncio
import time
import aiohttp
import configparser
import common_code


def output_exel(info,max_questions,dimensions,sheets,_type):

    if os.path.isfile('analysis.xlsx') is True :
        wb = load_workbook(filename = "analysis.xlsx")
        sheets = wb.get_sheet_names()
    else :
        wb1 = load_workbook(filename = "data.xlsx")
        wb = Workbook()
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        for sheet in sheets:
            wb.create_sheet(sheet)
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            ws1 = wb1.get_sheet_by_name(sheet.title)
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column = dimensions[sheet.title][1]+1)
            ws1 = wb1.get_sheet_by_name(sheet.title)

            for i in range(2,dimensions[sheet.title][1]+2) :
                ws.cell(row = 2,column = i,value = str(ws1.cell(row = 2, column = i).value))
            for i in range(1,dimensions[sheet.title][0]+3) :
                ws.cell(row = i,column = 1,value = str(ws1.cell(row = i, column = 1).value))
    if _type is 'sentiment' :
        key_value = 'score'
        tag_string = "sentiment analysis score : "
    if _type is 'keyword' :
        key_value = 'keyPhrases'
        tag_string = 'key words : '
    for j in range(max_questions):
        ans_no = 0
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            for i in range(dimensions[sheet.title][0]) :
                if dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value != None:
                    ws.cell(row = i+3, column = j+2, value = str(ws.cell(row = i+3, column = j+2).value) + str("\n") + tag_string + str(info[j][ans_no][key_value]))
                    ans_no += 1
                elif dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value == None:
                    ws.cell(row = i+3, column = j+2, value = str(tag_string + str(info[j][ans_no][key_value])))
                    ans_no += 1

    wb.save('analysis.xlsx')

def output_sentimnet_exel(info,max_questions,dimensions,sheets):

    if os.path.isfile('analysis.xlsx') is True :
        wb = load_workbook(filename = "analysis.xlsx")
        sheets = wb.get_sheet_names()
    else :
        wb = load_workbook(filename = "data.xlsx")
        wb.save("analysis.xlsx")
        wb = load_workbook(filename = "analysis.xlsx")

    for j in range(max_questions):
        ans_no = 0
        for sheet in wb :
            ws = wb.get_sheet_by_name(sheet.title)
            for i in range(dimensions[sheet.title][0]) :
                if dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value != None:
                    #if ws.cell(row = i+3, column = j+2).value == None or ws.cell(row = i+3, column = j+2).value == "" :
                        #info[j][ans_no]['sentiment']['score'] = 0

                    ws.cell(row = i+3, column = j+2, value = str(ws.cell(row = i+3, column = j+2).value) + str("\n") + 'sentiment score : ' + str(info[j][ans_no]['sentiment']['score']))
                    ans_no += 1
                elif dimensions[sheet.title][1] > j and ws.cell(row = i+3, column = j+2).value == None:
                    ws.cell(row = i+3, column = j+2, value = str('sentiment score : ' + str(info[j][ans_no]['sentiment']['score'])))
                    ans_no += 1

    wb.save('analysis.xlsx')


def generate_piechart_for_google_sentiment(sentiment) :
    for i in range(len(sentiment)) :
        a = [0,0,0]
        for sent in sentiment[i] :
            if sent['sentiment']['score'] > 1.5 :
                a[0] += 1
            elif sent['sentiment']['score'] < -1.5 :
                a[1] += 1
            else :
                a[2] += 1

        labels = 'positive', 'negative', 'neutral'
        explode = (0, 0.1,0.2)
        fig1, ax1 = plt.subplots()
        ax1.pie(a, explode=explode, labels=labels, autopct='%1.1f%%',shadow=True, startangle=90)
        ax1.axis('equal')

        plt.savefig('google_sentiment_pie_chart' + str(i+1) +".png")


@asyncio.coroutine
def call_url(url,text_data,i,request):
    req = {'document': {'type': 'PLAIN_TEXT','content': text_data,}}
    response = yield from aiohttp.ClientSession().post(url, data = json.dumps(req))
    response = yield from response.json()
    return [response,url,i]


def send_async_request(url,data) :
    async_res,res_info,temp_data = [],[],[]
    sentiment,keywords = [[]] * len(data),[[]] * len(data)
    regex = re.compile('[^a-zA-Z\s]')
    for q in range(len(data)):
        string = ""
        string1 = ""
        count = 0
        for i in data[q] :
            string1 += str(i) + ". "
            i = i.replace('.','')
            i = regex.sub("",i)
            if i != '' :
                string +=str(i) + ". "
            else :
                string += 'eat. '
        async_res.append(call_url(url[0],string,q,'Sentiment'))
        async_res.append(call_url(url[1],string1,q,'Keywords'))

    loop = asyncio.get_event_loop()
    res_info.append(loop.run_until_complete(asyncio.wait(async_res)))
    for i in range(len(res_info[0][0])) :
        val = res_info[0][0].pop().result()
        if val[1] == url[0] :
            sentiment[val[2]] = sentiment[val[2]] + val[0]['sentences']
        elif val[1] == url[1] :
            keywords[val[2]] = keywords[val[2]] + val[0]['tokens']

    return sentiment,keywords

def analysis_sentiment_and_keyword(data,max_questions,dimensions,sheets):
    url = ['https://language.googleapis.com/v1/documents:analyzeSentiment?key=YOUR_API_KEY','https://language.googleapis.com/v1/documents:analyzeSyntax?key=YOUR_API_KEY']
    sol,keys,no_of_sol,sol1,cloud_array,cloud_data = [],[],[],[],[],[]
    data,max_questions,dimensions,sheets = common_code.read_questions()
    tim = time.time()
    sentiment_data,syntax_data = send_async_request(url,data)
    print("time taken for keyword and sentiment analysis : ",time.time()-tim)
    generate_piechart_for_google_sentiment(sentiment_data)
    dictionary = [('ADJ','JJ'),('ADP','IN'),('ADV','RB'),('CONJ','CC'),('DET','DT'),('NOUN','NN'),('PRON','PRP'),('PRT','RP'),('VERB','VB'),('NUM','CD')]
    nltk_inp = []
    output_sentimnet_exel(sentiment_data,max_questions,dimensions,sheets)
    for i in range(max_questions) :
        nltk_inp.append([])
        cloud_array.append([])
        for val in syntax_data[i] :
            if val['partOfSpeech']['tag'] == 'NOUN' or val['partOfSpeech']['tag'] == 'VERB' :
                cloud_array[i].append(val['text']['content'])
            test = []
            test = [item[1] for item in dictionary if item[0] == val['partOfSpeech']['tag']]
            #if val['text']['content'] == 'between' :
                #print(val['partOfSpeech']['tag'])
            if test != [] :
                nltk_inp[i].append((val['text']['content'],test[0]))
    for i in range(len(cloud_array)) :
        text = ""
        for ans in cloud_array[i] :
            text = text + " " + ans
        cloud_data.append(text)


    common_code.generate_word_cloud(cloud_data,"google_keywords","google_thread")
